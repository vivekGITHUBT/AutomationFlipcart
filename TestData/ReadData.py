import openpyxl
wk= openpyxl.load_workbook("E://Test.xlsx")
#sh=wk["Sheet1"]
#row=sh.max_row
#print(row)
#cell=sh.cell(2,2)
#print(cell.value)
#It is running so we need to create methods

def Fetch_Number_Of_Rows(SheetName):
    sh=wk[SheetName]
    return sh.max_row
def Fetch_Cell_Data(SheetName,row,cell):
    sh=wk[SheetName]
    celldata=sh.cell(int(row),int(cell))
    return celldata.value
